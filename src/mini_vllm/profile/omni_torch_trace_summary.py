from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Optional


_STAGE_RE = re.compile(r"stage[_-](\d+)")
_RANK_RE = re.compile(r"ops_rank(\d+)\.xlsx$")

# Columns emitted by OmniTorchProfilerWrapper's "summary" sheet.
_NUMERIC_COLS = (
    "count",
    "self_cpu_time_total_us",
    "self_cuda_time_total_us",
    "self_cpu_memory_usage_bytes",
    "self_cuda_memory_usage_bytes",
)


def _parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Reduce vLLM-Omni per-worker torch-profiler artifacts (ops_rank*.xlsx "
            "under VLLM_TORCH_PROFILER_DIR) into a flat per-stage op summary, "
            "compatible with the jsonl/csv shape used by the other profilers."
        )
    )
    parser.add_argument("--trace_dir", type=str, required=True,
                        help="Directory passed as VLLM_TORCH_PROFILER_DIR / --torch_profiler_dir.")
    parser.add_argument("--out_jsonl", type=str, default="omni_ops_summary.jsonl")
    parser.add_argument("--out_csv", type=str, default="omni_ops_summary.csv")
    parser.add_argument("--top", type=int, default=25,
                        help="Top-N ops per (stage, rank) by self CUDA time.")
    return parser.parse_args(argv)


def _stage_from_path(session_dir: Path) -> Optional[int]:
    match = _STAGE_RE.search(session_dir.name)
    return int(match.group(1)) if match else None


def _rank_from_path(xlsx_path: Path) -> Optional[int]:
    match = _RANK_RE.search(xlsx_path.name)
    return int(match.group(1)) if match else None


def _read_summary_sheet(xlsx_path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "summary" not in wb.sheetnames:
            return []
        ws = wb["summary"]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return []
        records: list[dict[str, object]] = []
        for raw in rows_iter:
            record = {str(h): v for h, v in zip(header, raw)}
            records.append(record)
        return records
    finally:
        wb.close()


def _coerce(record: dict[str, object]) -> dict[str, object]:
    out: dict[str, object] = {"name": record.get("name")}
    for col in _NUMERIC_COLS:
        value = record.get(col)
        try:
            out[col] = float(value) if value is not None else None
        except (TypeError, ValueError):
            out[col] = None
    return out


def collect(trace_dir: str, top: int) -> list[dict[str, object]]:
    root = Path(trace_dir)
    if not root.exists():
        raise FileNotFoundError(trace_dir)

    rows: list[dict[str, object]] = []
    for xlsx_path in sorted(root.rglob("ops_rank*.xlsx")):
        session_dir = xlsx_path.parent
        stage = _stage_from_path(session_dir)
        rank = _rank_from_path(xlsx_path)
        records = [_coerce(r) for r in _read_summary_sheet(xlsx_path)]
        records.sort(
            key=lambda r: (r.get("self_cuda_time_total_us") or 0.0),
            reverse=True,
        )
        for rank_index, record in enumerate(records[:top]):
            rows.append(
                {
                    "record_type": "omni_op",
                    "session": session_dir.name,
                    "stage": stage,
                    "rank": rank,
                    "op_rank": rank_index,
                    **record,
                }
            )
    return rows


def _write_jsonl(path: str, rows: list[dict[str, object]]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row) + "\n")


def _write_csv(path: str, rows: list[dict[str, object]]) -> None:
    keys: list[str] = []
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)


def main(argv: Optional[list[str]] = None) -> int:
    args = _parse_args(argv)
    rows = collect(args.trace_dir, args.top)
    if not rows:
        print(
            f"No ops_rank*.xlsx found under {args.trace_dir}. "
            "Did the run set --torch_profiler_dir and complete the trace flush?"
        )
    _write_jsonl(args.out_jsonl, rows)
    _write_csv(args.out_csv, rows)
    print(f"Wrote {args.out_jsonl} and {args.out_csv} ({len(rows)} op rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
